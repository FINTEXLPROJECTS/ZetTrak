from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction
from django.http import HttpResponse
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.urls import reverse
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname
from .forms import CompanySignupForm, CustomAuthenticationForm, AdminEmployeeForm, EmployeeLoginForm, INDIAN_STATES
from .models import CustomUser, Company, EmployeeProfile, Attendance, LeaveRequest, LeaveType, UserLeaveBalance, Payroll, Department, Shift, Branch, Designation, RegularizationRequest, LeavePolicy, Notification, Announcement, NotificationSetting, PublicHoliday

def get_effective_days(start_date, end_date):
    """Returns number of days passed within the range [start_date, end_date] up to today."""
    today = date.today()
    if start_date > today:
        return 0
    effective_end = min(end_date, today)
    return (effective_end - start_date).days + 1

def create_notification(user, company, n_type, message):
    """Helper to create notification and check settings"""
    # Get or create settings for company
    settings, _ = NotificationSetting.objects.get_or_create(company=company)
    
    # Map notification type to setting field for EXTERNAL (Email/SMS)
    # For System (in-app), we usually want to allow most updates if system alerts are enabled
    type_map = {
        'Leave': settings.leave_request_notif,
        'Attendance': settings.attendance_alert_notif,
        'Payroll': settings.payroll_alert_notif,
        'Employee': settings.employee_update_notif,
        'System': settings.system_update_notif,
        'Announcement': True,
    }
    
    # Check if system notifications are enabled globally for the company
    if settings.enable_system:
        # Always allow status updates to the user about their own requests
        # Or if the specific type is enabled
        Notification.objects.create(
            user=user,
            company=company,
            notification_type=n_type,
            message=message
        )
        
    # Logic for Email/SMS/Push would go here based on settings.enable_email AND type_map
    if settings.enable_email and type_map.get(n_type, True):
        # send_email_logic(user, message)
        pass
import calendar as cal_module
from collections import defaultdict
from datetime import datetime, timedelta, date
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

@login_required
def attendance_list_admin(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    selected_date = request.GET.get('date', date.today().isoformat())
    attendances = Attendance.objects.filter(user__company=request.user.company, date=selected_date)
    
    # Get all employees to show who's missing
    employees = CustomUser.objects.filter(company=request.user.company, is_company_admin=False)
    
    context = {
        'attendances': attendances,
        'employees': employees,
        'selected_date': selected_date,
    }
    return render(request, 'zettrack_app/attendance_list_admin.html', context)

@login_required
def regularization_list_admin(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    requests = RegularizationRequest.objects.filter(user__company=request.user.company).order_by('-applied_on')
    return render(request, 'zettrack_app/regularization_list_admin.html', {'requests': requests})

@login_required
def approve_regularization(request, request_id, status):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    reg_request = RegularizationRequest.objects.get(id=request_id, user__company=request.user.company)
    reg_request.status = status
    reg_request.save()
    
    if status == 'approved':
        # Update attendance record
        attendance, created = Attendance.objects.get_or_create(
            user=reg_request.user,
            date=reg_request.attendance_date,
            defaults={'status': 'P'}
        )
        if reg_request.expected_punch_in:
            # Combine date and time
            attendance.punch_in = timezone.make_aware(timezone.datetime.combine(reg_request.attendance_date, reg_request.expected_punch_in))
        if reg_request.expected_punch_out:
            attendance.punch_out = timezone.make_aware(timezone.datetime.combine(reg_request.attendance_date, reg_request.expected_punch_out))
        attendance.save()
        
    # Notify Employee
    create_notification(
        user=reg_request.user,
        company=request.user.company,
        n_type='Attendance',
        message=f"Your regularization request for {reg_request.attendance_date} has been {status}."
    )
        
    messages.success(request, f"Request {status} successfully.")
    return redirect('regularization_list_admin')

@login_required
def attendance_calendar_employee(request):
    attendances = Attendance.objects.filter(user=request.user).order_by('-date')
    today = date.today()
    attendance = Attendance.objects.filter(user=request.user, date=today).first()
    return render(request, 'zettrack_app/attendance_calendar_employee.html', {'attendances': attendances, 'attendance': attendance})

@login_required
def regularization_request_view(request):
    if request.method == 'POST':
        attendance_date = request.POST.get('attendance_date')
        reason = request.POST.get('reason')
        punch_in = request.POST.get('punch_in')
        punch_out = request.POST.get('punch_out')
        
        RegularizationRequest.objects.create(
            user=request.user,
            attendance_date=attendance_date,
            reason=reason,
            expected_punch_in=punch_in if punch_in else None,
            expected_punch_out=punch_out if punch_out else None
        )
        admins = CustomUser.objects.filter(company=request.user.company, is_company_admin=True)
        for admin in admins:
            create_notification(
                user=admin,
                company=request.user.company,
                n_type='Attendance',
                message=f"{request.user.get_full_name()} submitted a regularization request for {attendance_date}."
            )
        messages.success(request, "Regularization request submitted.")
        return redirect('attendance_calendar_employee')
    return render(request, 'zettrack_app/regularization_request.html')

@login_required
def employee_list_view(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    employees = CustomUser.objects.filter(company=request.user.company, is_company_admin=False)
    return render(request, 'zettrack_app/employees_list.html', {'employees': employees})

@login_required
def add_employee_view(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    if request.method == 'POST':
        form = AdminEmployeeForm(request.POST, company=request.user.company)
        if form.is_valid():
            email = form.cleaned_data['email']
            phone = form.cleaned_data['phone']
            emp_code = form.cleaned_data['employee_code']
            
            if not email:
                email = f"{emp_code}@zettrak.com"
            
            # Check for existing user with same email or phone
            if CustomUser.objects.filter(email=email).exists():
                messages.error(request, f"Error: A user with email {email} already exists.")
            elif CustomUser.objects.filter(phone=phone).exists():
                messages.error(request, f"Error: A user with phone number {phone} already exists.")
            else:
                try:
                    user = CustomUser.objects.create_user(
                        email=email,
                        phone=phone,
                        first_name=form.cleaned_data['first_name'],
                        last_name=form.cleaned_data['last_name'],
                        company=request.user.company
                    )
                    if form.cleaned_data['password']:
                        user.set_password(form.cleaned_data['password'])
                    else:
                        user.set_password("ZetTrak123")
                    user.save()
                    
                    profile = form.save(commit=False)
                    profile.user = user
                    company = request.user.company
                    designation_name = form.cleaned_data.get('designation', '').strip()
                    if designation_name:
                        obj, _ = Designation.objects.get_or_create(company=company, name=designation_name)
                        profile.designation = obj
                    department_name = form.cleaned_data.get('department', '').strip()
                    if department_name:
                        obj, _ = Department.objects.get_or_create(company=company, name=department_name)
                        profile.department = obj
                    branch_name = form.cleaned_data.get('branch', '').strip()
                    if branch_name:
                        obj, _ = Branch.objects.get_or_create(company=company, name=branch_name)
                        profile.branch = obj
                    shift_name = form.cleaned_data.get('shift', '').strip()
                    if shift_name:
                        obj = Shift.objects.filter(company=company, name=shift_name).first()
                        if obj:
                            profile.shift = obj
                    profile.save()
                    
                    # Notify Admin
                    create_notification(
                        user=request.user,
                        company=request.user.company,
                        n_type='Employee',
                        message=f"New employee {user.get_full_name()} added to {profile.department.name if profile.department else 'N/A'} department."
                    )
                    
                    messages.success(request, f"Employee {user.get_full_name()} added successfully.")
                    return redirect('employee_list')
                except Exception as e:
                    messages.error(request, f"An error occurred: {str(e)}")
    else:
        form = AdminEmployeeForm(company=request.user.company)
    
    return render(request, 'zettrack_app/add_employee.html', {'form': form})

@login_required
def download_employee_template(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    # Secondary sheet for dropdown lists (can be hidden)
    data_sheet = wb.create_sheet(title="Lists")
    
    company = request.user.company
    depts = [d.name for d in Department.objects.filter(company=company)]
    desigs = [d.name for d in Designation.objects.filter(company=company)]
    branches = [b.name for b in Branch.objects.filter(company=company)]
    shifts = [s.name for s in Shift.objects.filter(company=company)]
    roles = ['Employee', 'Manager', 'HR/Admin']

    # Populate Lists sheet
    data_cols = [depts, desigs, branches, shifts, roles]
    data_names = ['Depts', 'Desigs', 'Branches', 'Shifts', 'Roles']
    list_ranges = {}

    for col, (name, values) in enumerate(zip(data_names, data_cols), start=1):
        data_sheet.cell(row=1, column=col, value=name).font = Font(bold=True)
        for row, val in enumerate(values, start=2):
            data_sheet.cell(row=row, column=col, value=val)
        
        if values:
            last_row = len(values) + 1
            col_letter = openpyxl.utils.get_column_letter(col)
            list_ranges[name] = f"={quote_sheetname('Lists')}!${col_letter}$2:${col_letter}${last_row}"
        else:
            list_ranges[name] = None

    headers = [
        'first_name', 'last_name', 'email', 'phone', 'employee_code',
        'department', 'designation', 'branch', 'shift', 'joining_date',
        'role', 'password'
    ]
    notes = [
        'Required', 'Required', 'Optional (auto-generated if blank)',
        'Required', 'Required', 'Optional (Use dropdown)', 'Optional (Use dropdown)', 'Optional (Use dropdown)',
        'Optional (Use dropdown)', 'Optional (YYYY-MM-DD)', 'Optional (Use dropdown)',
        'Optional (default: ZetTrak123)'
    ]

    header_fill = PatternFill(start_color='86BC25', end_color='86BC25', fill_type='solid')
    note_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    note_font = Font(italic=True, color='6B7280', size=9)
    border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB'),
    )

    for col, (header, note) in enumerate(zip(headers, notes), start=1):
        hcell = ws.cell(row=1, column=col, value=header)
        hcell.font = header_font
        hcell.fill = header_fill
        hcell.alignment = Alignment(horizontal='center', vertical='center')
        hcell.border = border

        ncell = ws.cell(row=2, column=col, value=note)
        ncell.font = note_font
        ncell.fill = note_fill
        ncell.alignment = Alignment(horizontal='center', wrap_text=True)
        ncell.border = border

        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

    # Apply Data Validations
    def apply_validation(col_name, range_ref):
        if range_ref:
            dv = DataValidation(type="list", formula1=range_ref, allow_blank=True)
            col_idx = headers.index(col_name) + 1
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            dv.add(f"{col_letter}3:{col_letter}1000")
            ws.add_data_validation(dv)

    apply_validation('department', list_ranges['Depts'])
    apply_validation('designation', list_ranges['Desigs'])
    apply_validation('branch', list_ranges['Branches'])
    apply_validation('shift', list_ranges['Shifts'])
    apply_validation('role', list_ranges['Roles'])

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30
    data_sheet.sheet_state = 'hidden'

    sample = ['John', 'Doe', 'john.doe@company.com', '9876543210', 'EMP001',
              depts[0] if depts else '', desigs[0] if desigs else '', 
              branches[0] if branches else '', shifts[0] if shifts else '',
              '2025-01-15', 'Employee', '']
    for col, val in enumerate(sample, start=1):
        cell = ws.cell(row=3, column=col, value=val)
        cell.border = border
        cell.alignment = Alignment(horizontal='left')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="employee_import_template.xlsx"'
    return response


@login_required
def import_employees(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Please select an Excel file to upload.')
            return redirect('import_employees')

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Invalid file type. Please upload an .xlsx or .xls file.')
            return redirect('import_employees')

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        except Exception:
            messages.error(request, 'Could not read the Excel file. Please use the provided template.')
            return redirect('import_employees')

        company = request.user.company
        created, skipped, errors = 0, 0, []

        # Row 1 = headers, Row 2 = notes, Row 3+ = data
        for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            if not any(row):
                continue

            def cell(i):
                v = row[i] if i < len(row) else None
                return str(v).strip() if v is not None else ''

            first_name = cell(0)
            last_name = cell(1)
            email = cell(2)
            phone = cell(3)
            emp_code = cell(4)
            department = cell(5)
            designation = cell(6)
            branch = cell(7)
            shift = cell(8)
            joining_date_raw = cell(9)
            role = cell(10) or 'Employee'
            password = cell(11) or 'ZetTrak123'

            if not first_name or not phone or not emp_code:
                errors.append(f"Row {row_num}: first_name, phone, and employee_code are required.")
                skipped += 1
                continue

            if not email:
                email = f"{emp_code.lower()}@zettrak.com"

            if CustomUser.objects.filter(email=email).exists():
                errors.append(f"Row {row_num}: Email '{email}' already exists — skipped.")
                skipped += 1
                continue

            if CustomUser.objects.filter(phone=phone).exists():
                errors.append(f"Row {row_num}: Phone '{phone}' already exists — skipped.")
                skipped += 1
                continue

            joining_date = None
            if joining_date_raw:
                from datetime import datetime as dt
                for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y'):
                    try:
                        joining_date = dt.strptime(joining_date_raw, fmt).date()
                        break
                    except ValueError:
                        pass

            try:
                with transaction.atomic():
                    user = CustomUser.objects.create_user(
                        email=email, phone=phone,
                        first_name=first_name, last_name=last_name,
                        company=company
                    )
                    user.set_password(password)
                    user.save()

                    profile = EmployeeProfile(user=user, employee_code=emp_code, role=role, joining_date=joining_date)
                    if department:
                        obj, _ = Department.objects.get_or_create(company=company, name=department)
                        profile.department = obj
                    
                    if designation:
                        obj, _ = Designation.objects.get_or_create(company=company, name=designation)
                        profile.designation = obj
                    if branch:
                        obj, _ = Branch.objects.get_or_create(company=company, name=branch)
                        profile.branch = obj
                    if shift:
                        obj = Shift.objects.filter(company=company, name=shift).first()
                        if obj:
                            profile.shift = obj
                    profile.save()
                created += 1
            except Exception as e:
                errors.append(f"Row {row_num}: Error — {str(e)}")
                skipped += 1

        if created:
            messages.success(request, f'Successfully imported {created} employee(s).')
        if skipped:
            messages.warning(request, f'{skipped} row(s) were skipped.')

        return render(request, 'zettrack_app/employee_import.html', {
            'done': True, 'created': created, 'skipped': skipped, 'errors': errors
        })

    return render(request, 'zettrack_app/employee_import.html', {'done': False})


@login_required
def attendance_list_view(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')

    selected_date_str = request.GET.get('date', date.today().isoformat())
    selected_date = date.fromisoformat(selected_date_str)

    month = int(request.GET.get('month', selected_date.month))
    year = int(request.GET.get('year', selected_date.year))

    company = request.user.company
    all_employees = CustomUser.objects.filter(company=company, is_company_admin=False).order_by('first_name')
    selected_employee_id = request.GET.get('employee_id', '')
    employees = all_employees
    if selected_employee_id:
        employees = all_employees.filter(id=selected_employee_id)

    # --- Daily attendance ---
    attendance_data = []
    for emp in employees:
        att = Attendance.objects.filter(user=emp, date=selected_date).first()
        work_hours = None
        if att and att.punch_in and att.punch_out:
            delta = att.punch_out - att.punch_in
            h = delta.seconds // 3600
            m = (delta.seconds % 3600) // 60
            work_hours = f"{h}h {m}m"
        attendance_data.append({
            'employee': emp,
            'attendance': att,
            'work_hours': work_hours,
        })

    # --- Monthly calendar ---
    month_start = date(year, month, 1)
    if month == 12:
        month_end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = date(year, month + 1, 1) - timedelta(days=1)

    month_attendances = Attendance.objects.filter(
        user__company=company, date__gte=month_start, date__lte=month_end
    )

    day_stats = {}
    today = date.today()
    
    if selected_employee_id:
        # Get attendance status per day for this specific employee
        emp_month_atts = month_attendances.filter(user_id=selected_employee_id)
        status_map = {att.date: att.status for att in emp_month_atts}
        
        for d in range(1, month_end.day + 1):
            curr_date = date(year, month, d)
            if curr_date.weekday() == 6:  # Sunday
                day_stats[d] = 'WO'
            elif curr_date > today:
                day_stats[d] = 'FUTURE'
            elif curr_date in status_map:
                st = status_map[curr_date]
                if st in ('P', 'HD'):
                    day_stats[d] = 'PRESENT'
                elif st == 'A':
                    day_stats[d] = 'ABSENT'
                elif st == 'L':
                    day_stats[d] = 'LEAVE'
                elif st == 'H':
                    day_stats[d] = 'HOLIDAY'
                else:
                    day_stats[d] = 'WO' if st == 'WO' else 'PRESENT'
            else:
                day_stats[d] = 'ABSENT'
    
    calendar_weeks = cal_module.monthcalendar(year, month)

    # --- Monthly report per employee ---
    # Update get_effective_days logic or handle it here to only count up to today and exclude Sundays?
    # Actually user said "completed days only counted", so effective_days is correct.
    # But if we want absent to NOT include Sundays:
    
    def get_working_days(start, end):
        count = 0
        curr = start
        today_obj = date.today()
        while curr <= end and curr <= today_obj:
            if curr.weekday() != 6: # Not Sunday
                count += 1
            curr += timedelta(days=1)
        return count

    working_days_count = get_working_days(month_start, month_end)

    emp_report = []
    for emp in all_employees: # Always show all employees in summary
        emp_atts = month_attendances.filter(user=emp)
        present = emp_atts.filter(status='P').count()
        half_day = emp_atts.filter(status='HD').count()
        on_leave = emp_atts.filter(status='L').count()
        
        # Total absent = (Working days passed) - (Present + Half Day + On Leave)
        # This counts days with no records AND days with explicit 'A' status as absent.
        absent = max(0, working_days_count - (present + half_day + on_leave))
        
        total_secs = sum(
            (a.punch_out - a.punch_in).seconds
            for a in emp_atts if a.punch_in and a.punch_out
        )
        th = total_secs // 3600
        tm = (total_secs % 3600) // 60
        emp_report.append({
            'employee': emp,
            'present': present,
            'half_day': half_day,
            'on_leave': on_leave,
            'absent': absent,
            'total_hours': f"{th}h {tm}m",
        })

    present_today = Attendance.objects.filter(
        user__company=company, date=date.today(), punch_in__isnull=False
    ).count()

    context = {
        'attendance_data': attendance_data,
        'selected_date': selected_date,
        'calendar_weeks': calendar_weeks,
        'day_stats': day_stats,
        'month': month,
        'year': year,
        'month_name': cal_module.month_name[month],
        'month_choices': [(i, cal_module.month_name[i]) for i in range(1, 13)],
        'emp_report': emp_report,
        'total_employees': all_employees.count(),
        'present_today': present_today,
        'today': date.today(),
        'all_employees': all_employees,
        'selected_employee_id': selected_employee_id,
    }
    return render(request, 'zettrack_app/attendance_list.html', context)

@login_required
def regularization_requests_view(request):
    if not request.user.is_company_admin:
        # For employees, show their own requests
        requests = RegularizationRequest.objects.filter(user=request.user).order_by('-applied_on')
        return render(request, 'zettrack_app/my_regularizations.html', {'requests': requests})
    
    # For admins, show all pending for their company
    requests = RegularizationRequest.objects.filter(user__company=request.user.company).order_by('-applied_on')
    return render(request, 'zettrack_app/admin_regularizations.html', {'requests': requests})

@login_required
def apply_regularization(request):
    if request.method == 'POST':
        attendance_date = request.POST.get('attendance_date')
        reason = request.POST.get('reason')
        in_time = request.POST.get('expected_punch_in')
        out_time = request.POST.get('expected_punch_out')
        
        RegularizationRequest.objects.create(
            user=request.user,
            attendance_date=attendance_date,
            reason=reason,
            expected_punch_in=in_time if in_time else None,
            expected_punch_out=out_time if out_time else None
        )
        admins = CustomUser.objects.filter(company=request.user.company, is_company_admin=True)
        for admin in admins:
            create_notification(
                user=admin,
                company=request.user.company,
                n_type='Attendance',
                message=f"{request.user.get_full_name()} submitted a regularization request for {attendance_date}."
            )
        messages.success(request, "Regularization request submitted.")
        return redirect('employee_dashboard')
    return render(request, 'zettrack_app/apply_regularization.html')

@login_required
def employee_detail_view(request, employee_id):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    employee = CustomUser.objects.get(id=employee_id, company=request.user.company)
    return render(request, 'zettrack_app/employee_detail.html', {'employee': employee})

@login_required
def edit_employee_view(request, employee_id):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    employee = CustomUser.objects.get(id=employee_id, company=request.user.company)
    profile = employee.employee_profile
    
    if request.method == 'POST':
        form = AdminEmployeeForm(request.POST, instance=profile, company=request.user.company)
        if form.is_valid():
            employee.first_name = form.cleaned_data['first_name']
            employee.last_name = form.cleaned_data['last_name']
            
            email = form.cleaned_data['email']
            phone = form.cleaned_data['phone']
            
            # Check for existing user with same email or phone (excluding current user)
            if email and CustomUser.objects.filter(email=email).exclude(id=employee.id).exists():
                messages.error(request, f"Error: A user with email {email} already exists.")
            elif CustomUser.objects.filter(phone=phone).exclude(id=employee.id).exists():
                messages.error(request, f"Error: A user with phone number {phone} already exists.")
            else:
                if email:
                    employee.email = email
                employee.phone = phone
                
                if form.cleaned_data['password']:
                    employee.set_password(form.cleaned_data['password'])
                
                employee.save()
                form.save()
                
                messages.success(request, f"Employee {employee.get_full_name()} updated successfully.")
                return redirect('employee_list')
    else:
        initial_data = {
            'first_name': employee.first_name,
            'last_name': employee.last_name,
            'email': employee.email,
            'phone': employee.phone,
        }
        form = AdminEmployeeForm(instance=profile, initial=initial_data, company=request.user.company)
    
    return render(request, 'zettrack_app/edit_employee.html', {'form': form, 'employee': employee})


@login_required
def delete_employee_view(request, employee_id):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    employee = get_object_or_404(CustomUser, id=employee_id, company=request.user.company, is_company_admin=False)
    if request.method == 'POST':
        name = employee.get_full_name()
        employee.delete()
        messages.success(request, f"Employee '{name}' deleted successfully.")
    return redirect('employee_list')

import random
from django.utils import timezone
from datetime import date

@login_required
def dashboard_redirect(request):
    if request.user.is_company_admin:
        return redirect('admin_dashboard')
    else:
        return redirect('employee_dashboard')

@login_required
def admin_dashboard(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')

    company = request.user.company
    employees = CustomUser.objects.filter(company=company, is_company_admin=False)
    total_employees = employees.count()

    today = date.today()
    present_today = Attendance.objects.filter(user__company=company, date=today, punch_in__isnull=False).count()
    leave_request_user_ids = set(LeaveRequest.objects.filter(
        user__company=company, status='approved', start_date__lte=today, end_date__gte=today
    ).values_list('user_id', flat=True))
    attendance_leave_user_ids = set(Attendance.objects.filter(
        user__company=company, date=today, status='L'
    ).values_list('user_id', flat=True))
    on_leave_user_ids = leave_request_user_ids | attendance_leave_user_ids
    on_leave = len(on_leave_user_ids)
    absent = total_employees - present_today - on_leave

    pending_leaves = LeaveRequest.objects.filter(user__company=company, status='pending').count()

    # Late arrivals count
    today_attendances = Attendance.objects.filter(
        user__company=company, date=today, punch_in__isnull=False, status='P'
    ).select_related('user__employee_profile__shift')
    from datetime import time as dtime
    DEFAULT_LATE_TIME = dtime(11, 0)  # 11:00 AM fallback
    late_count = 0
    for att in today_attendances:
        try:
            shift = att.user.employee_profile.shift
            cutoff = shift.start_time if shift else DEFAULT_LATE_TIME
            if timezone.localtime(att.punch_in).time() > cutoff:
                late_count += 1
        except Exception:
            pass

    context = {
        'total_employees': total_employees,
        'present_today': present_today,
        'on_leave': on_leave,
        'absent': absent,
        'pending_leaves': pending_leaves,
        'company': company,
        'late_count': late_count,
    }
    return render(request, 'zettrack_app/admin_dashboard.html', context)


@login_required
def admin_late_arrivals(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')

    selected_date = request.GET.get('date', date.today().isoformat())

    attendances = Attendance.objects.filter(
        user__company=request.user.company,
        date=selected_date,
        punch_in__isnull=False,
        status='P'
    ).select_related('user', 'user__employee_profile__shift')

    from datetime import datetime as dt, time as dtime
    DEFAULT_LATE_TIME = dtime(11, 0)  # 11:00 AM fallback

    late_records = []
    for att in attendances:
        try:
            if not att.punch_in:
                continue
            punch_in_local = timezone.localtime(att.punch_in)
            punch_in_time = punch_in_local.time()
            d = date.fromisoformat(selected_date)

            shift = att.user.employee_profile.shift
            if shift:
                cutoff = shift.start_time
            else:
                cutoff = DEFAULT_LATE_TIME

            if punch_in_time > cutoff:
                late_by_secs = (
                    dt.combine(d, punch_in_time) - dt.combine(d, cutoff)
                ).total_seconds()
                late_records.append({
                    'user': att.user,
                    'punch_in': punch_in_local,
                    'shift': shift,
                    'late_minutes': int(late_by_secs // 60),
                })
        except Exception:
            pass

    context = {
        'late_records': late_records,
        'selected_date': selected_date,
    }
    return render(request, 'zettrack_app/admin_late_arrivals.html', context)


@login_required
def admin_present_today(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')

    selected_date = request.GET.get('date', date.today().isoformat())

    attendances = Attendance.objects.filter(
        user__company=request.user.company,
        date=selected_date,
        punch_in__isnull=False,
    ).select_related('user', 'user__employee_profile__shift',
                     'user__employee_profile__department',
                     'user__employee_profile__designation')

    records = []
    for att in attendances:
        punch_in_local = timezone.localtime(att.punch_in)
        punch_out_local = timezone.localtime(att.punch_out) if att.punch_out else None
        work_hours = None
        if att.punch_out:
            delta = att.punch_out - att.punch_in
            h = delta.seconds // 3600
            m = (delta.seconds % 3600) // 60
            work_hours = f"{h}h {m}m"
        records.append({
            'user': att.user,
            'punch_in': punch_in_local,
            'punch_out': punch_out_local,
            'work_hours': work_hours,
            'status': att.status,
        })

    context = {
        'records': records,
        'selected_date': selected_date,
    }
    return render(request, 'zettrack_app/admin_present_today.html', context)


@login_required
def admin_absent_today(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')

    selected_date_str = request.GET.get('date', date.today().isoformat())
    selected_date_obj = date.fromisoformat(selected_date_str)

    company = request.user.company
    all_employees = CustomUser.objects.filter(company=company, is_company_admin=False).select_related(
        'employee_profile__department', 'employee_profile__designation'
    )

    # Employees who punched in on selected date
    present_ids = set(Attendance.objects.filter(
        user__company=company, date=selected_date_obj, punch_in__isnull=False
    ).values_list('user_id', flat=True))

    # Employees on approved leave on selected date
    on_leave_ids = set(LeaveRequest.objects.filter(
        user__company=company, status='approved',
        start_date__lte=selected_date_obj, end_date__gte=selected_date_obj
    ).values_list('user_id', flat=True))

    absent_employees = [
        emp for emp in all_employees
        if emp.id not in present_ids and emp.id not in on_leave_ids
    ]

    context = {
        'absent_employees': absent_employees,
        'selected_date': selected_date_str,
        'absent_count': len(absent_employees),
    }
    return render(request, 'zettrack_app/admin_absent_today.html', context)


@login_required
def admin_on_leave_today(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')

    selected_date_str = request.GET.get('date', date.today().isoformat())
    selected_date_obj = date.fromisoformat(selected_date_str)

    company = request.user.company
    leave_requests = LeaveRequest.objects.filter(
        user__company=company,
        status='approved',
        start_date__lte=selected_date_obj,
        end_date__gte=selected_date_obj,
    ).select_related('user', 'user__employee_profile__department',
                     'user__employee_profile__designation')

    # Also include employees manually marked as Leave in attendance (no LeaveRequest)
    leave_request_user_ids = set(leave_requests.values_list('user_id', flat=True))
    manual_leave_attendances = Attendance.objects.filter(
        user__company=company,
        date=selected_date_obj,
        status='L',
    ).exclude(user_id__in=leave_request_user_ids).select_related(
        'user', 'user__employee_profile__department',
        'user__employee_profile__designation'
    )

    total_on_leave = leave_requests.count() + manual_leave_attendances.count()
    context = {
        'leave_requests': leave_requests,
        'manual_leave_attendances': manual_leave_attendances,
        'total_on_leave': total_on_leave,
        'selected_date': selected_date_str,
    }
    return render(request, 'zettrack_app/admin_on_leave_today.html', context)


@login_required
def employee_holidays(request):
    if request.user.is_company_admin:
        return redirect('admin_dashboard')
    today = date.today()
    holidays = PublicHoliday.objects.filter(company=request.user.company).order_by('date')
    return render(request, 'zettrack_app/employee_holidays.html', {
        'holidays': holidays,
        'today': today,
    })

@login_required
def employee_profile(request):
    if request.user.is_company_admin:
        return redirect('admin_dashboard')
    try:
        profile = request.user.employee_profile
    except Exception:
        profile = None
    holidays = PublicHoliday.objects.filter(company=request.user.company).order_by('date')
    return render(request, 'zettrack_app/employee_profile.html', {'profile': profile, 'holidays': holidays})

@login_required
def employee_dashboard(request):
    if request.user.is_company_admin:
        return redirect('admin_dashboard')
    
    today = date.today()
    attendance = Attendance.objects.filter(user=request.user, date=today).first()
    leave_balance = 15 # Mock data
    upcoming_holidays = [] # Mock data
    
    recent_leaves = LeaveRequest.objects.filter(user=request.user).order_by('-applied_on')[:5]
    
    context = {
        'attendance': attendance,
        'leave_balance': leave_balance,
        'recent_leaves': recent_leaves,
    }
    return render(request, 'zettrack_app/employee_dashboard.html', context)

@login_required
def punch_in(request):
    today = date.today()
    attendance, created = Attendance.objects.get_or_create(user=request.user, date=today)
    if not attendance.punch_in:
        attendance.punch_in = timezone.now()
        attendance.save()
        messages.success(request, "Punched in successfully.")
    return redirect('attendance_calendar_employee')

@login_required
def punch_out(request):
    today = date.today()
    attendance = Attendance.objects.filter(user=request.user, date=today).first()
    if attendance and not attendance.punch_out:
        attendance.punch_out = timezone.now()
        attendance.save()
        messages.success(request, "Punched out successfully.")
    return redirect('attendance_calendar_employee')

@login_required
def leave_apply(request):
    if request.method == 'POST':
        leave_type = request.POST.get('leave_type')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        reason = request.POST.get('reason')
        
        LeaveRequest.objects.create(
            user=request.user,
            leave_type=leave_type,
            start_date=start_date,
            end_date=end_date,
            reason=reason
        )
        
        # Notify Admin
        admins = CustomUser.objects.filter(company=request.user.company, is_company_admin=True)
        for admin in admins:
            create_notification(
                user=admin,
                company=request.user.company,
                n_type='Leave',
                message=f"{request.user.get_full_name()} applied for leave from {start_date} to {end_date}"
            )
            
        messages.success(request, "Leave request submitted.")
        return redirect('employee_dashboard')
    return render(request, 'zettrack_app/leave_apply.html')

@login_required
def admin_payroll_dashboard(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    company = request.user.company
    today = date.today()
    
    # Get payrolls for current month
    month_start = today.replace(day=1)
    current_payrolls = Payroll.objects.filter(user__company=company, month=month_start)
    
    # Summary stats
    total_payroll = sum(p.net_salary for p in current_payrolls)
    total_employees = CustomUser.objects.filter(company=company, is_company_admin=False).count()
    processed_count = current_payrolls.filter(status='Processed').count()
    pending_count = total_employees - processed_count
    
    # For Salary Structure (just a list of employees for now)
    employees = CustomUser.objects.filter(company=company, is_company_admin=False).select_related('employee_profile')
    
    # Payroll History
    payroll_history = Payroll.objects.filter(user__company=company).order_by('-month', 'user__first_name')
    
    context = {
        'total_payroll': total_payroll,
        'total_employees': total_employees,
        'processed_count': processed_count,
        'pending_count': pending_count,
        'employees': employees,
        'payroll_history': payroll_history,
        'current_month': month_start,
    }
    return render(request, 'zettrack_app/admin_payroll_dashboard.html', context)

@login_required
def payslips_view(request):
    payrolls = Payroll.objects.filter(user=request.user).order_by('-month')
    return render(request, 'zettrack_app/payslips.html', {'payrolls': payrolls})

def index(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'zettrack_app/home.html')

def signup_view(request):
    if request.method == 'POST':
        form = CompanySignupForm(request.POST, request.FILES)
        if form.is_valid():
            # Create Company
            company = form.save()
            
            # Create User (Admin)
            user = CustomUser.objects.create_user(
                email=form.cleaned_data['email'],
                password=form.cleaned_data['password'],
                first_name=form.cleaned_data['full_name'],
                phone=form.cleaned_data['phone'],
                company=company,
                is_company_admin=True
            )
            
            login(request, user)
            messages.success(request, f"Registration successful for {company.name}.")
            return redirect('dashboard')
        else:
            messages.error(request, "Unsuccessful registration. Please correct the errors.")
    else:
        form = CompanySignupForm(initial={'country': 'India'})
    
    return render(request, 'zettrack_app/signup.html', {'form': form})

def login_view(request):
    if request.method == 'POST':
        login_type = request.POST.get('login_type', 'admin')
        
        if login_type == 'employee':
            emp_id = request.POST.get('employee_id', '').strip()
            phone = request.POST.get('phone', '').strip()
            
            try:
                # Search by phone number first as it is unique in CustomUser
                user = CustomUser.objects.filter(phone=phone).first()
                if user and hasattr(user, 'employee_profile'):
                    profile = user.employee_profile
                    # Compare employee_code case-insensitively
                    if profile.employee_code.strip().lower() == emp_id.lower():
                        # Manually set the backend for login as it hasn't been through authenticate()
                        if not hasattr(user, 'backend'):
                            user.backend = 'django.contrib.auth.backends.ModelBackend'
                        login(request, user)
                        messages.info(request, f"Welcome back, {user.first_name}.")
                        return redirect('dashboard')
                    else:
                        messages.error(request, "Invalid Employee ID or Mobile Number.")
                else:
                    messages.error(request, "Invalid Employee ID or Mobile Number.")
            except Exception as e:
                messages.error(request, f"An error occurred: {str(e)}")
        else:
            form = CustomAuthenticationForm(request, data=request.POST)
            if form.is_valid():
                email = form.cleaned_data.get('username')
                password = form.cleaned_data.get('password')
                user = authenticate(email=email, password=password)
                if user is not None:
                    login(request, user)
                    messages.info(request, f"You are now logged in as {user.first_name}.")
                    return redirect('dashboard')
                else:
                    messages.error(request,"Invalid email or password.")
            else:
                messages.error(request,"Invalid email or password.")
    
    # Forms for initial render
    admin_form = CustomAuthenticationForm()
    employee_form = EmployeeLoginForm()
    
    return render(request, 'zettrack_app/login.html', {
        'form': admin_form, # Keep for backward compatibility or simple case
        'admin_form': admin_form,
        'employee_form': employee_form
    })

def logout_view(request):
    logout(request)
    messages.info(request, "You have successfully logged out.")
    return redirect('index')

@login_required
def mark_attendance_view(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        att_date = request.POST.get('att_date')
        punch_in_str = request.POST.get('punch_in')
        punch_out_str = request.POST.get('punch_out')
        status = request.POST.get('status', 'P')
        try:
            emp = CustomUser.objects.get(id=user_id, company=request.user.company)
            att, _ = Attendance.objects.get_or_create(user=emp, date=att_date)
            if punch_in_str:
                att.punch_in = timezone.make_aware(datetime.strptime(f"{att_date} {punch_in_str}", "%Y-%m-%d %H:%M"))
            if punch_out_str:
                att.punch_out = timezone.make_aware(datetime.strptime(f"{att_date} {punch_out_str}", "%Y-%m-%d %H:%M"))
            att.status = status
            att.save()
            messages.success(request, f"Attendance marked for {emp.get_full_name()}.")
        except Exception as e:
            messages.error(request, f"Error: {str(e)}")
    return redirect(f"{reverse('admin_attendance')}?date={att_date}")

def features_view(request):
    return render(request, 'zettrack_app/features.html')

def about_view(request):
    return render(request, 'zettrack_app/about.html')

def contact_view(request):
    return render(request, 'zettrack_app/contact.html')

def how_to_use_view(request):
    return render(request, 'zettrack_app/how_to_use.html')

def pricing_view(request):
    return render(request, 'zettrack_app/pricing.html')

@login_required
def admin_leave_dashboard(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    today = date.today()
    company = request.user.company
    
    # Filters for sections
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    dept_id = request.GET.get('department')
    emp_id = request.GET.get('employee')
    type_id = request.GET.get('leave_type')
    status = request.GET.get('status')
    today_only = request.GET.get('today_only') == '1'
    
    # Leave requests for today (always based on company and today's date)
    requests_applied_today = LeaveRequest.objects.filter(user__company=company, applied_on__date=today).count()
    
    pending_approvals = LeaveRequest.objects.filter(user__company=company, status='pending').count()
    approved_leaves = LeaveRequest.objects.filter(user__company=company, status='approved').count()
    rejected_leaves = LeaveRequest.objects.filter(user__company=company, status='rejected').count()
    
    # Employees on leave today
    on_leave_today = LeaveRequest.objects.filter(
        user__company=company, 
        status='approved', 
        start_date__lte=today, 
        end_date__gte=today
    )
    if dept_id:
        on_leave_today = on_leave_today.filter(user__employee_profile__department_id=dept_id)
    if emp_id:
        on_leave_today = on_leave_today.filter(user_id=emp_id)
        
    on_leave_today_count = on_leave_today.count()
    
    # Statistics for Chart (Last 6 months)
    chart_labels = []
    chart_data = []
    for i in range(5, -1, -1):
        d = today - timedelta(days=i*30)
        m = d.month
        y = d.year
        chart_labels.append(cal_module.month_name[m])
        chart_qs = LeaveRequest.objects.filter(user__company=company, status='approved', start_date__year=y, start_date__month=m)
        if dept_id:
            chart_qs = chart_qs.filter(user__employee_profile__department_id=dept_id)
        if emp_id:
            chart_qs = chart_qs.filter(user_id=emp_id)
        chart_data.append(chart_qs.count())
    
    # All Leave Requests (Tab: Requests)
    all_requests = LeaveRequest.objects.filter(user__company=company).order_by('-applied_on')
    if dept_id:
        all_requests = all_requests.filter(user__employee_profile__department_id=dept_id)
    if emp_id:
        all_requests = all_requests.filter(user_id=emp_id)
    if type_id:
        all_requests = all_requests.filter(leave_type_obj_id=type_id)
    if status:
        all_requests = all_requests.filter(status=status)
    if today_only:
        all_requests = all_requests.filter(applied_on__date=today)

    for req in all_requests:
        req.duration = (req.end_date - req.start_date).days + 1
    
    # All Leave Types
    all_leave_types = LeaveType.objects.filter(company=company)
    
    # Leave Policy
    leave_policy, created = LeavePolicy.objects.get_or_create(company=company)
    if request.method == 'POST' and request.POST.get('form_type') == 'policy':
        leave_policy.max_leaves_per_year = request.POST.get('max_leaves', 30)
        leave_policy.accrual_policy = request.POST.get('accrual', 'Monthly')
        leave_policy.carry_forward_limit = request.POST.get('cf_limit', 5)
        leave_policy.approval_workflow = request.POST.get('workflow', 'Single Level')
        leave_policy.save()
        messages.success(request, "Leave policy updated successfully.")
        return redirect('/dashboard/admin/leaves/?tab=policies')

    # Initializing Leave Balances
    employees = CustomUser.objects.filter(company=company, is_company_admin=False)
    for emp in employees:
        for lt in all_leave_types:
            UserLeaveBalance.objects.get_or_create(user=emp, leave_type=lt, year=today.year, defaults={'total_days': lt.total_days})
            
    # Leave Balances (Tab: Balance)
    all_balances = UserLeaveBalance.objects.filter(user__company=company, year=today.year)
    if dept_id:
        all_balances = all_balances.filter(user__employee_profile__department_id=dept_id)
    if emp_id:
        all_balances = all_balances.filter(user_id=emp_id)
    if type_id:
        all_balances = all_balances.filter(leave_type_id=type_id)
        
    for bal in all_balances:
        used = LeaveRequest.objects.filter(user=bal.user, leave_type_obj=bal.leave_type, status='approved', start_date__year=today.year)
        bal.used_days = sum((r.end_date - r.start_date).days + 1 for r in used)
        bal.remaining = bal.total_days - bal.used_days
        bal.save()

    # --- Leave Calendar Logic ---
    calendar_weeks = cal_module.monthcalendar(year, month)
    month_start = date(year, month, 1)
    if month == 12: month_end = date(year + 1, 1, 1) - timedelta(days=1)
    else: month_end = date(year, month + 1, 1) - timedelta(days=1)
    
    calendar_leaves = LeaveRequest.objects.filter(
        user__company=company, status='approved',
        start_date__lte=month_end, end_date__gte=month_start
    )
    if dept_id:
        calendar_leaves = calendar_leaves.filter(user__employee_profile__department_id=dept_id)
    if emp_id:
        calendar_leaves = calendar_leaves.filter(user_id=emp_id)
    
    # Map leaves to days and prepare week data
    day_leaves = defaultdict(list)
    for lreq in calendar_leaves:
        curr = max(lreq.start_date, month_start)
        last = min(lreq.end_date, month_end)
        while curr <= last:
            day_leaves[curr.day].append(lreq)
            curr += timedelta(days=1)
            
    final_weeks = []
    for week in calendar_weeks:
        week_days = []
        for d in week:
            if d == 0:
                week_days.append({'day': 0, 'leaves': []})
            else:
                week_days.append({'day': d, 'leaves': day_leaves[d]})
        final_weeks.append(week_days)

    # Prepare report data grouped by user (Tab: Reports)
    report_by_user = []
    report_employees = employees
    if dept_id:
        report_employees = report_employees.filter(employee_profile__department_id=dept_id)
    if emp_id:
        report_employees = report_employees.filter(id=emp_id)
        
    for emp in report_employees:
        emp_balances = UserLeaveBalance.objects.filter(user=emp, year=today.year)
        total_allowed = sum(b.total_days for b in emp_balances)
        total_used = sum(b.used_days for b in emp_balances)
        total_pending = LeaveRequest.objects.filter(user=emp, status='pending', start_date__year=today.year).count()
        total_rejected = LeaveRequest.objects.filter(user=emp, status='rejected', start_date__year=today.year).count()
        utilization = int((total_used / total_allowed * 100) if total_allowed > 0 else 0)
        
        report_by_user.append({
            'user': emp,
            'total_used': total_used,
            'total_pending': total_pending,
            'total_rejected': total_rejected,
            'utilization': utilization
        })

    context = {
        'requests_applied_today': requests_applied_today,
        'pending_approvals': pending_approvals,
        'approved_leaves': approved_leaves,
        'rejected_leaves': rejected_leaves,
        'on_leave_today': on_leave_today,
        'on_leave_today_count': on_leave_today_count,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'all_requests': all_requests,
        'all_leave_types': all_leave_types,
        'all_balances': all_balances,
        'final_weeks': final_weeks,
        'report_by_user': report_by_user,
        'leave_policy': leave_policy,
        'month': month,
        'year': year,
        'month_name': cal_module.month_name[month],
        'departments': Department.objects.filter(company=company),
        'all_employees': employees,
        'selected_dept': int(dept_id) if dept_id else None,
        'selected_emp': int(emp_id) if emp_id else None,
        'selected_type': int(type_id) if type_id else None,
        'selected_status': status,
        'today_only': today_only,
        'today': today,
        'month_choices': [(i, cal_module.month_name[i]) for i in range(1, 13)],
        'year_choices': range(today.year - 1, today.year + 2),
    }
    return render(request, 'zettrack_app/admin_leave_dashboard.html', context)

@login_required
def add_leave_type(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        total_days = request.POST.get('total_days', 12)
        carry_forward = request.POST.get('carry_forward') == 'on'
        description = request.POST.get('description', '')
        
        LeaveType.objects.create(
            company=request.user.company,
            name=name,
            total_days=total_days,
            carry_forward=carry_forward,
            description=description
        )
        messages.success(request, f"Leave type '{name}' created successfully.")
    return redirect('admin_leaves')

@login_required
def edit_leave_type(request, type_id):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    leave_type = LeaveType.objects.get(id=type_id, company=request.user.company)
    if request.method == 'POST':
        leave_type.name = request.POST.get('name')
        leave_type.total_days = request.POST.get('total_days', 12)
        leave_type.carry_forward = request.POST.get('carry_forward') == 'on'
        leave_type.description = request.POST.get('description', '')
        leave_type.save()
        messages.success(request, f"Leave type '{leave_type.name}' updated.")
    return redirect('admin_leaves')

@login_required
def delete_leave_type(request, type_id):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    leave_type = LeaveType.objects.get(id=type_id, company=request.user.company)
    name = leave_type.name
    leave_type.delete()
    messages.success(request, f"Leave type '{name}' deleted.")
    return redirect('admin_leaves')

@login_required
def adjust_leave_balance(request, bal_id):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    bal = UserLeaveBalance.objects.get(id=bal_id, user__company=request.user.company)
    if request.method == 'POST':
        new_total = request.POST.get('total_days')
        bal.total_days = new_total
        bal.save()
        messages.success(request, f"Balance adjusted for {bal.user.get_full_name()}.")
    return redirect('admin_leaves')

@login_required
def add_extra_leave(request, bal_id):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    bal = UserLeaveBalance.objects.get(id=bal_id, user__company=request.user.company)
    if request.method == 'POST':
        extra = request.POST.get('extra_days', 0)
        bal.total_days = float(bal.total_days) + float(extra)
        bal.save()
        messages.success(request, f"Added {extra} days to {bal.user.get_full_name()}.")
    return redirect('admin_leaves')

@login_required
def reset_leave_balance(request, bal_id):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    bal = UserLeaveBalance.objects.get(id=bal_id, user__company=request.user.company)
    bal.total_days = bal.leave_type.total_days
    bal.save()
    messages.success(request, f"Balance reset for {bal.user.get_full_name()}.")
    return redirect('admin_leaves')

@login_required
def approve_leave(request, leave_id):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    leave = LeaveRequest.objects.get(id=leave_id, user__company=request.user.company)
    leave.status = 'approved'
    leave.save()
    
    # Create attendance records for the leave period
    current_date = leave.start_date
    while current_date <= leave.end_date:
        Attendance.objects.update_or_create(
            user=leave.user,
            date=current_date,
            defaults={'status': 'L'}
        )
        current_date += timedelta(days=1)
        
    # Notify Employee
    create_notification(
        user=leave.user,
        company=request.user.company,
        n_type='Leave',
        message=f"Your leave from {leave.start_date} to {leave.end_date} has been approved."
    )
    
    messages.success(request, f"Leave for {leave.user.get_full_name()} approved.")
    return redirect('admin_leaves')

@login_required
def reject_leave(request, leave_id):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    leave = LeaveRequest.objects.get(id=leave_id, user__company=request.user.company)
    leave.status = 'rejected'
    leave.save()
    
    # Notify Employee
    create_notification(
        user=leave.user,
        company=request.user.company,
        n_type='Leave',
        message=f"Your leave request for {leave.start_date} has been rejected."
    )
    
    messages.success(request, f"Leave for {leave.user.get_full_name()} rejected.")
    return redirect('admin_leaves')

def employee_signup_view(request):
    if request.method == 'POST':
        form = EmployeeSignupForm(request.POST)
        if form.is_valid():
            # In a real scenario, we'd find the user by OTP or invite code
            # For this MVP, we'll assume they are setting up their profile after an invite.
            messages.success(request, "Employee profile setup successful.")
            return redirect('index')
    else:
        form = EmployeeSignupForm()
    return render(request, 'zettrack_app/employee_signup.html', {'form': form})

@login_required
def export_attendance_excel(request):
    if not request.user.is_company_admin:
        return HttpResponse("Unauthorized", status=401)
    
    month = int(request.GET.get('month', date.today().month))
    year = int(request.GET.get('year', date.today().year))
    
    company = request.user.company
    employees = CustomUser.objects.filter(company=company, is_company_admin=False)
    
    month_start = date(year, month, 1)
    if month == 12:
        month_end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = date(year, month + 1, 1) - timedelta(days=1)
        
    month_attendances = Attendance.objects.filter(
        user__company=company, date__gte=month_start, date__lte=month_end
    )
    
    effective_days = get_effective_days(month_start, month_end)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance_{cal_module.month_name[month]}_{year}"
    
    headers = ["Employee Name", "Employee Code", "Present", "Absent", "Half Day", "On Leave", "Total Hours"]
    ws.append(headers)
    
    for emp in employees:
        emp_atts = month_attendances.filter(user=emp)
        present = emp_atts.filter(status='P').count()
        half_day = emp_atts.filter(status='HD').count()
        on_leave = emp_atts.filter(status='L').count()
        absent = max(0, effective_days - emp_atts.count())
        total_secs = sum(
            (a.punch_out - a.punch_in).seconds
            for a in emp_atts if a.punch_in and a.punch_out
        )
        th = total_secs // 3600
        tm = (total_secs % 3600) // 60
        
        emp_code = "-"
        if hasattr(emp, 'employee_profile'):
            emp_code = emp.employee_profile.employee_code or "-"
            
        ws.append([
            emp.get_full_name(),
            emp_code,
            present,
            absent,
            half_day,
            on_leave,
            f"{th}h {tm}m"
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Attendance_Report_{cal_module.month_name[month]}_{year}.xlsx'
    wb.save(response)
    return response

@login_required
def export_attendance_pdf(request):
    if not request.user.is_company_admin:
        return HttpResponse("Unauthorized", status=401)
    
    month = int(request.GET.get('month', date.today().month))
    year = int(request.GET.get('year', date.today().year))
    
    company = request.user.company
    employees = CustomUser.objects.filter(company=company, is_company_admin=False)
    
    month_start = date(year, month, 1)
    if month == 12:
        month_end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = date(year, month + 1, 1) - timedelta(days=1)
        
    month_attendances = Attendance.objects.filter(
        user__company=company, date__gte=month_start, date__lte=month_end
    )
    
    effective_days = get_effective_days(month_start, month_end)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    
    styles = getSampleStyleSheet()
    title = f"Attendance Report: {cal_module.month_name[month]} {year} - {company.name}"
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 12))
    
    data = [["Employee", "Emp ID", "Present", "Absent", "Half Day", "Leave", "Work Hours"]]
    for emp in employees:
        emp_atts = month_attendances.filter(user=emp)
        present = emp_atts.filter(status='P').count()
        half_day = emp_atts.filter(status='HD').count()
        on_leave = emp_atts.filter(status='L').count()
        absent = max(0, effective_days - emp_atts.count())
        total_secs = sum(
            (a.punch_out - a.punch_in).seconds
            for a in emp_atts if a.punch_in and a.punch_out
        )
        th = total_secs // 3600
        tm = (total_secs % 3600) // 60
        
        emp_code = "-"
        if hasattr(emp, 'employee_profile'):
            emp_code = emp.employee_profile.employee_code or "-"
            
        data.append([
            emp.get_full_name(),
            emp_code,
            str(present),
            str(absent),
            str(half_day),
            str(on_leave),
            f"{th}h {tm}m"
        ])
        
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Attendance_Report_{cal_module.month_name[month]}_{year}.pdf'
    return response

@login_required
def admin_notifications(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    
    company = request.user.company
    
    # Summary stats for dashboard
    # (These would normally be real counts from a model)
    stats = {
        'total_sent': 150,
        'unread_alerts': 12,
        'active_announcements': 3,
        'system_health': 'Good'
    }
    
    return render(request, 'zettrack_app/admin_notifications.html', {
        'stats': stats,
    })

@login_required
def export_payroll_excel(request):
    if not request.user.is_company_admin:
        return HttpResponse("Unauthorized", status=401)
    
    today = date.today()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    month_start = date(year, month, 1)
    
    company = request.user.company
    payrolls = Payroll.objects.filter(user__company=company, month=month_start)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Payroll_{cal_module.month_name[month]}_{year}"
    
    headers = ["Employee Name", "Emp ID", "Basic Salary", "Allowances", "Deductions", "Net Salary", "Status"]
    ws.append(headers)
    
    for p in payrolls:
        emp_code = "-"
        if hasattr(p.user, 'employee_profile'):
            emp_code = p.user.employee_profile.employee_code or "-"
        ws.append([
            p.user.get_full_name(),
            emp_code,
            float(p.basic_salary),
            float(p.allowances),
            float(p.deductions),
            float(p.net_salary),
            p.status
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Payroll_Report_{cal_module.month_name[month]}_{year}.xlsx'
    wb.save(response)
    return response

@login_required
def export_payroll_pdf(request):
    if not request.user.is_company_admin:
        return HttpResponse("Unauthorized", status=401)
    
    today = date.today()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    month_start = date(year, month, 1)
    
    company = request.user.company
    payrolls = Payroll.objects.filter(user__company=company, month=month_start)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()
    
    title = f"Payroll Report: {cal_module.month_name[month]} {year} - {company.name}"
    elements.append(Paragraph(title, styles['Title']))
    elements.append(Spacer(1, 12))
    
    data = [["Employee", "Emp ID", "Basic", "Allowances", "Deductions", "Net Salary", "Status"]]
    for p in payrolls:
        emp_code = "-"
        if hasattr(p.user, 'employee_profile'):
            emp_code = p.user.employee_profile.employee_code or "-"
        data.append([
            p.user.get_full_name(),
            emp_code,
            str(p.basic_salary),
            str(p.allowances),
            str(p.deductions),
            str(p.net_salary),
            p.status
        ])
        
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Payroll_Report_{cal_module.month_name[month]}_{year}.pdf'
    return response

@login_required
def export_payslip_pdf(request, payroll_id):
    payroll = Payroll.objects.get(id=payroll_id)
    if not request.user.is_company_admin and payroll.user != request.user:
        return HttpResponse("Unauthorized", status=401)
        
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    # Header
    elements.append(Paragraph(f"<b>{payroll.user.company.name}</b>", styles['Title']))
    elements.append(Paragraph("Monthly Salary Slip", styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Employee Info
    emp_code = "-"
    if hasattr(payroll.user, 'employee_profile'):
        emp_code = payroll.user.employee_profile.employee_code or "-"
        
    info_data = [
        ["Employee Name:", payroll.user.get_full_name(), "Emp ID:", emp_code],
        ["Pay Period:", payroll.month.strftime('%B %Y'), "Date Generated:", date.today().strftime('%d %b %Y')],
    ]
    info_table = Table(info_data, colWidths=[100, 150, 100, 150])
    info_table.setStyle(TableStyle([('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'), ('ALIGN', (0,0), (-1, -1), 'LEFT')]))
    elements.append(info_table)
    elements.append(Spacer(1, 24))
    
    # Salary Breakdown
    salary_data = [
        ["Description", "Earnings", "Description", "Deductions"],
        ["Basic Salary", f"₹{payroll.basic_salary}", "Professional Tax", "₹200.00"],
        ["Allowances", f"₹{payroll.allowances}", "PF/ESI", str(payroll.deductions - 200)],
        ["", "", "Total Deductions", f"₹{payroll.deductions}"],
        ["Gross Earnings", f"₹{payroll.basic_salary + payroll.allowances}", "Net Salary", f"₹{payroll.net_salary}"],
    ]
    salary_table = Table(salary_data, colWidths=[150, 100, 150, 100])
    salary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
    ]))
    elements.append(salary_table)
    
    elements.append(Spacer(1, 48))
    elements.append(Paragraph("This is a computer-generated document and does not require a signature.", styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Payslip_{payroll.user.first_name}_{payroll.month.strftime("%b_%Y")}.pdf'
    return response

# --- Department Views ---

@login_required
def department_list_view(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    departments = Department.objects.filter(company=request.user.company)
    dept_data = []
    for dept in departments:
        count = EmployeeProfile.objects.filter(department=dept).count()
        dept_data.append({'department': dept, 'employee_count': count})
    return render(request, 'zettrack_app/department_list.html', {'dept_data': dept_data})


@login_required
def add_department_view(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            if Department.objects.filter(company=request.user.company, name__iexact=name).exists():
                messages.error(request, 'A department with this name already exists.')
            else:
                Department.objects.create(company=request.user.company, name=name)
                messages.success(request, f'Department "{name}" added successfully.')
                return redirect('department_list')
        else:
            messages.error(request, 'Department name is required.')
    return render(request, 'zettrack_app/add_department.html', {'dept': None})


@login_required
def edit_department_view(request, dept_id):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    dept = Department.objects.get(id=dept_id, company=request.user.company)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            dept.name = name
            dept.save()
            messages.success(request, 'Department updated successfully.')
            return redirect('department_list')
        else:
            messages.error(request, 'Department name is required.')
    return render(request, 'zettrack_app/add_department.html', {'dept': dept})


@login_required
def delete_department_view(request, dept_id):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    dept = Department.objects.get(id=dept_id, company=request.user.company)
    if request.method == 'POST':
        dept.delete()
        messages.success(request, 'Department deleted successfully.')
    return redirect('department_list')


@login_required
def department_employees_view(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    departments = Department.objects.filter(company=request.user.company)
    selected_dept_id = request.GET.get('dept')
    selected_dept = None
    employees = CustomUser.objects.none()
    if selected_dept_id:
        selected_dept = Department.objects.filter(id=selected_dept_id, company=request.user.company).first()
        if selected_dept:
            employees = CustomUser.objects.filter(
                company=request.user.company,
                is_company_admin=False,
                employee_profile__department=selected_dept
            )
    else:
        employees = CustomUser.objects.filter(company=request.user.company, is_company_admin=False)
    return render(request, 'zettrack_app/department_employees.html', {
        'departments': departments,
        'selected_dept': selected_dept,
        'employees': employees,
        'selected_dept_id': selected_dept_id,
    })


@login_required
def department_reports_view(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    today = date.today()
    month_start = date(today.year, today.month, 1)
    departments = Department.objects.filter(company=request.user.company)
    report_data = []
    for dept in departments:
        emp_profiles = EmployeeProfile.objects.filter(department=dept)
        emp_users = [p.user for p in emp_profiles]
        count = len(emp_users)
        att_this_month = Attendance.objects.filter(
            user__in=emp_users,
            date__gte=month_start,
            date__lte=today,
            status='P'
        ).count()
        working_days = (today - month_start).days + 1
        total_possible = count * working_days
        rate = round((att_this_month / total_possible * 100), 1) if total_possible > 0 else 0
        report_data.append({
            'department': dept,
            'employee_count': count,
            'present_days': att_this_month,
            'attendance_rate': rate,
        })
    return render(request, 'zettrack_app/department_reports.html', {
        'report_data': report_data,
        'month_name': today.strftime('%B %Y'),
    })


# --- Designation Views ---

@login_required
def designation_list_view(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    designations = Designation.objects.filter(company=request.user.company)
    desig_data = []
    for desig in designations:
        count = EmployeeProfile.objects.filter(designation=desig).count()
        desig_data.append({'designation': desig, 'employee_count': count})
    return render(request, 'zettrack_app/designation_list.html', {'desig_data': desig_data})


@login_required
def add_designation_view(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            if Designation.objects.filter(company=request.user.company, name__iexact=name).exists():
                messages.error(request, 'A designation with this name already exists.')
            else:
                Designation.objects.create(company=request.user.company, name=name)
                messages.success(request, f'Designation "{name}" added successfully.')
                return redirect('designation_list')
        else:
            messages.error(request, 'Designation name is required.')
    return render(request, 'zettrack_app/add_designation.html', {'desig': None})


@login_required
def edit_designation_view(request, desig_id):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    desig = Designation.objects.get(id=desig_id, company=request.user.company)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            desig.name = name
            desig.save()
            messages.success(request, 'Designation updated successfully.')
            return redirect('designation_list')
        else:
            messages.error(request, 'Designation name is required.')
    return render(request, 'zettrack_app/add_designation.html', {'desig': desig})


@login_required
def delete_designation_view(request, desig_id):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    desig = Designation.objects.get(id=desig_id, company=request.user.company)
    if request.method == 'POST':
        desig.delete()
        messages.success(request, 'Designation deleted successfully.')
    return redirect('designation_list')


@login_required
def designation_employees_view(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    designations = Designation.objects.filter(company=request.user.company)
    selected_desig_id = request.GET.get('desig')
    selected_desig = None
    employees = CustomUser.objects.none()
    if selected_desig_id:
        selected_desig = Designation.objects.filter(id=selected_desig_id, company=request.user.company).first()
        if selected_desig:
            employees = CustomUser.objects.filter(
                company=request.user.company,
                is_company_admin=False,
                employee_profile__designation=selected_desig
            )
    else:
        employees = CustomUser.objects.filter(company=request.user.company, is_company_admin=False)
    return render(request, 'zettrack_app/designation_employees.html', {
        'designations': designations,
        'selected_desig': selected_desig,
        'employees': employees,
        'selected_desig_id': selected_desig_id,
    })


@login_required
def designation_hierarchy_view(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    designations = Designation.objects.filter(company=request.user.company)
    hierarchy = []
    for desig in designations:
        profiles = EmployeeProfile.objects.filter(designation=desig).select_related('department', 'user')
        departments = {}
        for p in profiles:
            dept_name = p.department.name if p.department else 'Unassigned'
            departments.setdefault(dept_name, []).append(p.user)
        hierarchy.append({
            'designation': desig,
            'employee_count': profiles.count(),
            'departments': departments,
        })
    return render(request, 'zettrack_app/designation_hierarchy.html', {'hierarchy': hierarchy})


@login_required
def designation_reports_view(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    today = date.today()
    month_start = date(today.year, today.month, 1)
    designations = Designation.objects.filter(company=request.user.company)
    report_data = []
    for desig in designations:
        emp_profiles = EmployeeProfile.objects.filter(designation=desig)
        emp_users = [p.user for p in emp_profiles]
        count = len(emp_users)
        att_this_month = Attendance.objects.filter(
            user__in=emp_users,
            date__gte=month_start,
            date__lte=today,
            status='P'
        ).count()
        working_days = (today - month_start).days + 1
        total_possible = count * working_days
        rate = round((att_this_month / total_possible * 100), 1) if total_possible > 0 else 0
        report_data.append({
            'designation': desig,
            'employee_count': count,
            'present_days': att_this_month,
            'attendance_rate': rate,
        })
    return render(request, 'zettrack_app/designation_reports.html', {
        'report_data': report_data,
        'month_name': today.strftime('%B %Y'),
    })


# --- Reports Section ---

@login_required
def report_attendance(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    company = request.user.company
    today = date.today()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    dept_id = request.GET.get('dept')

    month_start = date(year, month, 1)
    month_end = date(year, month + 1, 1) - timedelta(days=1) if month < 12 else date(year + 1, 1, 1) - timedelta(days=1)
    total_days = (month_end - month_start).days + 1

    employees = CustomUser.objects.filter(company=company, is_company_admin=False)
    if dept_id:
        employees = employees.filter(employee_profile__department_id=dept_id)

    month_attendances = Attendance.objects.filter(user__in=employees, date__gte=month_start, date__lte=month_end)

    effective_days = get_effective_days(month_start, month_end)
    report_data = []
    for emp in employees:
        emp_atts = month_attendances.filter(user=emp)
        present = emp_atts.filter(status='P').count()
        half_day = emp_atts.filter(status='HD').count()
        on_leave = emp_atts.filter(status='L').count()
        absent = max(0, effective_days - emp_atts.count())
        total_secs = sum((a.punch_out - a.punch_in).seconds for a in emp_atts if a.punch_in and a.punch_out)
        th, tm = total_secs // 3600, (total_secs % 3600) // 60
        report_data.append({
            'employee': emp,
            'present': present,
            'half_day': half_day,
            'on_leave': on_leave,
            'absent': absent,
            'total_hours': f"{th}h {tm}m",
        })

    return render(request, 'zettrack_app/report_attendance.html', {
        'report_data': report_data,
        'month': month,
        'year': year,
        'month_name': cal_module.month_name[month],
        'month_choices': [(i, cal_module.month_name[i]) for i in range(1, 13)],
        'year_choices': range(today.year - 1, today.year + 2),
        'departments': Department.objects.filter(company=company),
        'selected_dept': dept_id,
        'total_days': effective_days,
    })


@login_required
def report_payroll(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    company = request.user.company
    today = date.today()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    month_start = date(year, month, 1)

    payrolls = Payroll.objects.filter(user__company=company, month=month_start).select_related('user', 'user__employee_profile')
    total_net = sum(p.net_salary for p in payrolls)
    total_basic = sum(p.basic_salary for p in payrolls)
    total_deductions = sum(p.deductions for p in payrolls)

    return render(request, 'zettrack_app/report_payroll.html', {
        'payrolls': payrolls,
        'month': month,
        'year': year,
        'month_name': cal_module.month_name[month],
        'month_choices': [(i, cal_module.month_name[i]) for i in range(1, 13)],
        'year_choices': range(today.year - 1, today.year + 2),
        'total_net': total_net,
        'total_basic': total_basic,
        'total_deductions': total_deductions,
    })


@login_required
def report_employees(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    company = request.user.company
    dept_id = request.GET.get('dept')
    desig_id = request.GET.get('desig')
    status = request.GET.get('status')

    employees = CustomUser.objects.filter(company=company, is_company_admin=False).select_related('employee_profile')
    if dept_id:
        employees = employees.filter(employee_profile__department_id=dept_id)
    if desig_id:
        employees = employees.filter(employee_profile__designation_id=desig_id)
    if status == 'active':
        employees = employees.filter(is_active=True)
    elif status == 'inactive':
        employees = employees.filter(is_active=False)

    return render(request, 'zettrack_app/report_employees.html', {
        'employees': employees,
        'departments': Department.objects.filter(company=company),
        'designations': Designation.objects.filter(company=company),
        'selected_dept': dept_id,
        'selected_desig': desig_id,
        'selected_status': status,
        'total_count': employees.count(),
    })


@login_required
def report_leaves(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    company = request.user.company
    today = date.today()
    year = int(request.GET.get('year', today.year))
    dept_id = request.GET.get('dept')

    employees = CustomUser.objects.filter(company=company, is_company_admin=False)
    if dept_id:
        employees = employees.filter(employee_profile__department_id=dept_id)

    leave_types = LeaveType.objects.filter(company=company)
    report_data = []
    for emp in employees:
        balances = []
        for lt in leave_types:
            bal = UserLeaveBalance.objects.filter(user=emp, leave_type=lt, year=year).first()
            used = LeaveRequest.objects.filter(user=emp, leave_type_obj=lt, status='approved', start_date__year=year)
            used_days = sum((r.end_date - r.start_date).days + 1 for r in used)
            total = bal.total_days if bal else lt.total_days
            balances.append({'leave_type': lt, 'total': total, 'used': used_days, 'remaining': float(total) - used_days})
        report_data.append({'employee': emp, 'balances': balances})

    return render(request, 'zettrack_app/report_leaves.html', {
        'report_data': report_data,
        'leave_types': leave_types,
        'year': year,
        'year_choices': range(today.year - 1, today.year + 2),
        'departments': Department.objects.filter(company=company),
        'selected_dept': dept_id,
    })


@login_required
def report_departments(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    company = request.user.company
    today = date.today()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    month_start = date(year, month, 1)
    month_end = date(year, month + 1, 1) - timedelta(days=1) if month < 12 else date(year + 1, 1, 1) - timedelta(days=1)
    working_days = (month_end - month_start).days + 1

    departments = Department.objects.filter(company=company)
    report_data = []
    for dept in departments:
        emp_profiles = EmployeeProfile.objects.filter(department=dept)
        emp_users = [p.user for p in emp_profiles]
        count = len(emp_users)
        present = Attendance.objects.filter(user__in=emp_users, date__gte=month_start, date__lte=month_end, status='P').count()
        on_leave = Attendance.objects.filter(user__in=emp_users, date__gte=month_start, date__lte=month_end, status='L').count()
        total_possible = count * working_days
        rate = round((present / total_possible * 100), 1) if total_possible > 0 else 0
        report_data.append({
            'department': dept,
            'employee_count': count,
            'present_days': present,
            'leave_days': on_leave,
            'attendance_rate': rate,
        })

    return render(request, 'zettrack_app/report_departments.html', {
        'report_data': report_data,
        'month': month,
        'year': year,
        'month_name': cal_module.month_name[month],
        'month_choices': [(i, cal_module.month_name[i]) for i in range(1, 13)],
        'year_choices': range(today.year - 1, today.year + 2),
        'working_days': working_days,
    })


@login_required
def report_custom(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    company = request.user.company
    today = date.today()
    report_data = []
    applied = False

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    dept_id = request.GET.get('dept')
    desig_id = request.GET.get('desig')
    status_filter = request.GET.get('status')

    if start_date_str and end_date_str:
        applied = True
        start_date = date.fromisoformat(start_date_str)
        end_date = date.fromisoformat(end_date_str)
        effective_days = get_effective_days(start_date, end_date)

        employees = CustomUser.objects.filter(company=company, is_company_admin=False)
        if dept_id:
            employees = employees.filter(employee_profile__department_id=dept_id)
        if desig_id:
            employees = employees.filter(employee_profile__designation_id=desig_id)
        if status_filter == 'active':
            employees = employees.filter(is_active=True)
        elif status_filter == 'inactive':
            employees = employees.filter(is_active=False)

        attendances = Attendance.objects.filter(user__in=employees, date__gte=start_date, date__lte=end_date)
        for emp in employees:
            emp_atts = attendances.filter(user=emp)
            present = emp_atts.filter(status='P').count()
            half_day = emp_atts.filter(status='HD').count()
            on_leave = emp_atts.filter(status='L').count()
            absent = max(0, effective_days - emp_atts.count())
            total_secs = sum((a.punch_out - a.punch_in).seconds for a in emp_atts if a.punch_in and a.punch_out)
            th, tm = total_secs // 3600, (total_secs % 3600) // 60
            report_data.append({
                'employee': emp,
                'present': present,
                'half_day': half_day,
                'on_leave': on_leave,
                'absent': absent,
                'total_hours': f"{th}h {tm}m",
            })

    return render(request, 'zettrack_app/report_custom.html', {
        'report_data': report_data,
        'applied': applied,
        'departments': Department.objects.filter(company=company),
        'designations': Designation.objects.filter(company=company),
        'start_date': start_date_str or '',
        'end_date': end_date_str or '',
        'selected_dept': dept_id,
        'selected_desig': desig_id,
        'selected_status': status_filter,
        'today': today.isoformat(),
    })

@login_required
def notification_dashboard(request):
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    notifications = Notification.objects.filter(user=request.user)
    
    # Filters
    n_type = request.GET.get('type')
    date_filter = request.GET.get('date') # Expected YYYY-MM
    
    if n_type:
        notifications = notifications.filter(notification_type=n_type)
    
    if date_filter:
        try:
            year, month = map(int, date_filter.split('-'))
            notifications = notifications.filter(created_at__year=year, created_at__month=month)
        except ValueError:
            pass
            
    context = {
        'notifications': notifications,
        'types': ['Leave', 'Attendance', 'Payroll', 'Employee', 'System', 'Announcement'],
        'selected_type': n_type,
        'selected_date': date_filter,
    }
    return render(request, 'zettrack_app/notifications.html', context)

@login_required
def mark_notification_read(request, notif_id):
    notif = Notification.objects.get(id=notif_id, user=request.user)
    notif.is_read = True
    notif.save()
    return redirect('notification_dashboard')

@login_required
def mark_all_notifications_read(request):
    Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
    messages.success(request, "All notifications marked as read.")
    return redirect('notification_dashboard')

@login_required
def get_unread_notifications(request):
    """AJAX endpoint to get unread notifications for real-time alerts"""
    unread = Notification.objects.filter(user=request.user, is_read=False).order_by('-created_at')
    data = []
    for n in unread:
        data.append({
            'id': n.id,
            'message': n.message,
            'type': n.notification_type,
            'created_at': n.created_at.strftime('%h:%i %p')
        })
    return JsonResponse({'notifications': data, 'count': len(data)})

@login_required
def delete_notification(request, notif_id):
    notif = Notification.objects.get(id=notif_id, user=request.user)
    notif.delete()
    return redirect('notification_dashboard')

@login_required
def notification_settings_view(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
        
    settings, created = NotificationSetting.objects.get_or_create(company=request.user.company)
    
    if request.method == 'POST':
        settings.enable_email = 'enable_email' in request.POST
        settings.enable_sms = 'enable_sms' in request.POST
        settings.enable_push = 'enable_push' in request.POST
        settings.enable_system = 'enable_system' in request.POST
        
        settings.leave_request_notif = 'leave_request_notif' in request.POST
        settings.payroll_alert_notif = 'payroll_alert_notif' in request.POST
        settings.attendance_alert_notif = 'attendance_alert_notif' in request.POST
        settings.employee_update_notif = 'employee_update_notif' in request.POST
        settings.system_update_notif = 'system_update_notif' in request.POST
        
        settings.save()
        messages.success(request, "Notification settings updated.")
        return redirect('notification_settings')
        
    return render(request, 'zettrack_app/notification_settings.html', {'settings': settings})

@login_required
def send_announcement(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
        
    if request.method == 'POST':
        title = request.POST.get('title')
        message = request.POST.get('message')
        
        announcement = Announcement.objects.create(
            company=request.user.company,
            title=title,
            message=message,
            created_by=request.user
        )
        
        # Notify all employees
        employees = CustomUser.objects.filter(company=request.user.company)
        for emp in employees:
            create_notification(
                user=emp,
                company=request.user.company,
                n_type='Announcement',
                message=f"New Announcement: {title}"
            )
            
        messages.success(request, "Announcement sent successfully.")
        return redirect('admin_notifications')
        
    return render(request, 'zettrack_app/send_announcement.html')

@login_required
def process_payroll_all(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
        
    company = request.user.company
    today = date.today()
    month_start = today.replace(day=1)
    
    employees = CustomUser.objects.filter(company=company, is_company_admin=False)
    
    for emp in employees:
        # Mock payroll generation logic
        payroll, created = Payroll.objects.get_or_create(
            user=emp,
            month=month_start,
            defaults={
                'basic_salary': 30000,
                'allowances': 5000,
                'deductions': 2000,
                'net_salary': 33000,
                'status': 'Processed'
            }
        )
        
        if created:
            create_notification(
                user=emp,
                company=company,
                n_type='Payroll',
                message=f"Payroll processed for {month_start.strftime('%B %Y')}. Your payslip is ready."
            )
            
    # Notify admin
    admin = CustomUser.objects.filter(company=company, is_company_admin=True).first()
    if admin:
        create_notification(
            user=admin,
            company=company,
            n_type='Payroll',
            message=f"Payroll generation completed for {month_start.strftime('%B %Y')}."
        )
        
    messages.success(request, "Payroll processed and notifications sent.")
    return redirect('admin_payroll_dashboard')

@login_required
def admin_profile_edit(request):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    user = request.user
    company = user.company
    if request.method == 'POST':
        user.first_name = request.POST.get('first_name', '').strip()
        user.last_name = request.POST.get('last_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        if phone and phone != user.phone:
            if CustomUser.objects.filter(phone=phone).exclude(pk=user.pk).exists():
                messages.error(request, "This phone number is already in use.")
                return redirect('admin_profile_edit')
            user.phone = phone
        new_password = request.POST.get('new_password', '').strip()
        if new_password:
            user.set_password(new_password)
        user.save()
        company.name = request.POST.get('company_name', '').strip() or company.name
        company.industry = request.POST.get('industry', '').strip()
        company.state = request.POST.get('state', '').strip()
        company.city = request.POST.get('city', '').strip()
        if 'company_logo' in request.FILES:
            company.logo = request.FILES['company_logo']
        company.save()
        messages.success(request, "Profile updated successfully.")
        return redirect('admin_profile_edit')
    return render(request, 'zettrack_app/admin_profile_edit.html', {'user': user, 'company': company, 'states': INDIAN_STATES})

@login_required
def admin_edit_regularization(request, request_id):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    reg = RegularizationRequest.objects.get(id=request_id, user__company=request.user.company)
    if request.method == 'POST':
        reg.attendance_date = request.POST.get('attendance_date', reg.attendance_date)
        reg.reason = request.POST.get('reason', reg.reason)
        punch_in = request.POST.get('expected_punch_in', '').strip()
        punch_out = request.POST.get('expected_punch_out', '').strip()
        reg.expected_punch_in = punch_in if punch_in else None
        reg.expected_punch_out = punch_out if punch_out else None
        reg.status = request.POST.get('status', reg.status)
        reg.save()
        create_notification(
            user=reg.user,
            company=reg.user.company,
            n_type='Attendance',
            message=f"Your regularization request for {reg.attendance_date} has been {reg.status}."
        )
        messages.success(request, "Regularization request updated.")
        return redirect('regularization_list_admin')
    return render(request, 'zettrack_app/admin_edit_regularization.html', {'reg': reg})

@login_required
def admin_edit_attendance(request, employee_id, attendance_date):
    if not request.user.is_company_admin:
        return redirect('employee_dashboard')
    employee = CustomUser.objects.get(id=employee_id, company=request.user.company)
    attendance, _ = Attendance.objects.get_or_create(
        user=employee,
        date=attendance_date,
        defaults={'status': 'P'}
    )
    if request.method == 'POST':
        status = request.POST.get('status', attendance.status)
        punch_in_str = request.POST.get('punch_in', '').strip()
        punch_out_str = request.POST.get('punch_out', '').strip()
        attendance.status = status
        if punch_in_str:
            from django.utils import timezone as tz
            naive_in = datetime.strptime(f"{attendance_date} {punch_in_str}", "%Y-%m-%d %H:%M")
            attendance.punch_in = tz.make_aware(naive_in)
        else:
            attendance.punch_in = None
        if punch_out_str:
            from django.utils import timezone as tz
            naive_out = datetime.strptime(f"{attendance_date} {punch_out_str}", "%Y-%m-%d %H:%M")
            attendance.punch_out = tz.make_aware(naive_out)
        else:
            attendance.punch_out = None
        attendance.save()
        messages.success(request, f"Attendance updated for {employee.get_full_name()}.")
        return redirect(f"/dashboard/admin/attendance/?date={attendance_date}")
    return render(request, 'zettrack_app/admin_edit_attendance.html', {
        'employee': employee,
        'attendance': attendance,
        'attendance_date': attendance_date,
    })
